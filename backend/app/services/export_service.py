"""
Export Service
Handles Excel and PDF export functionality
"""
from typing import List, Optional
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

from app.models.surat_masuk import SuratMasuk
from app.models.surat_keluar import SuratKeluar


class ExportService:
    """Service for exporting data to various formats"""
    
    @staticmethod
    def export_surat_masuk_to_excel(surat_list: List[SuratMasuk]) -> BytesIO:
        """
        Export Surat Masuk to Excel
        Returns BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Surat Masuk"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "No", "Nomor Surat", "Tanggal Surat", "Tanggal Terima",
            "Pengirim", "Perihal", "Kategori", "Status", "Prioritas"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        for idx, surat in enumerate(surat_list, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=surat.nomor_surat)
            ws.cell(row=idx, column=3, value=surat.tanggal_surat.strftime("%Y-%m-%d") if surat.tanggal_surat else "")
            ws.cell(row=idx, column=4, value=surat.tanggal_terima.strftime("%Y-%m-%d") if surat.tanggal_terima else "")
            ws.cell(row=idx, column=5, value=surat.pengirim or "")
            ws.cell(row=idx, column=6, value=surat.perihal or "")
            ws.cell(row=idx, column=7, value=surat.kategori.nama if surat.kategori else "")
            ws.cell(row=idx, column=8, value=surat.status.value if surat.status else "")
            ws.cell(row=idx, column=9, value=surat.priority.value if surat.priority else "")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def export_surat_keluar_to_excel(surat_list: List[SuratKeluar]) -> BytesIO:
        """
        Export Surat Keluar to Excel
        Returns BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Surat Keluar"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "No", "Nomor Surat", "Tanggal Surat", "Penerima",
            "Perihal", "Kategori", "Status", "Prioritas"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        for idx, surat in enumerate(surat_list, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=surat.nomor_surat_keluar)
            ws.cell(row=idx, column=3, value=surat.tanggal_surat.strftime("%Y-%m-%d") if surat.tanggal_surat else "")
            ws.cell(row=idx, column=4, value=surat.penerima or "")
            ws.cell(row=idx, column=5, value=surat.perihal or "")
            ws.cell(row=idx, column=6, value=surat.kategori.nama if surat.kategori else "")
            ws.cell(row=idx, column=7, value=surat.status.value if surat.status else "")
            ws.cell(row=idx, column=8, value=surat.priority.value if surat.priority else "")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def export_surat_masuk_to_pdf(surat_list: List[SuratMasuk], title: str = "Laporan Surat Masuk") -> BytesIO:
        """
        Export Surat Masuk to PDF
        Returns BytesIO buffer containing the PDF file
        """
        buffer = BytesIO()
        
        # Create PDF with landscape orientation
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.5 * cm))
        
        # Table data
        data = [
            ["No", "Nomor Surat", "Tgl Surat", "Tgl Terima", "Pengirim", "Perihal", "Kategori", "Status"]
        ]
        
        for idx, surat in enumerate(surat_list, start=1):
            data.append([
                str(idx),
                surat.nomor_surat,
                surat.tanggal_surat.strftime("%Y-%m-%d") if surat.tanggal_surat else "",
                surat.tanggal_terima.strftime("%Y-%m-%d") if surat.tanggal_terima else "",
                surat.pengirim[:20] if surat.pengirim else "",
                surat.perihal[:30] if surat.perihal else "",
                surat.kategori.nama if surat.kategori else "",
                surat.status.value if surat.status else ""
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def export_surat_keluar_to_pdf(surat_list: List[SuratKeluar], title: str = "Laporan Surat Keluar") -> BytesIO:
        """
        Export Surat Keluar to PDF
        Returns BytesIO buffer containing the PDF file
        """
        buffer = BytesIO()
        
        # Create PDF with landscape orientation
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.5 * cm))
        
        # Table data
        data = [
            ["No", "Nomor Surat", "Tgl Surat", "Penerima", "Perihal", "Kategori", "Status"]
        ]
        
        for idx, surat in enumerate(surat_list, start=1):
            data.append([
                str(idx),
                surat.nomor_surat_keluar,
                surat.tanggal_surat.strftime("%Y-%m-%d") if surat.tanggal_surat else "",
                surat.penerima[:25] if surat.penerima else "",
                surat.perihal[:35] if surat.perihal else "",
                surat.kategori.nama if surat.kategori else "",
                surat.status.value if surat.status else ""
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
